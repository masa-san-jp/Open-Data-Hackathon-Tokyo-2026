#!/usr/bin/env python3
"""まちかどQRの候補地点と安全に表示できる周辺地点を生成する。

実行:
    python3 scripts/build_points.py

重要:
- 帰宅支援ステーションは「設置候補」であり、設置済み台帳ではない。
- 疑わしい座標を補正・推定しない。除外又は隔離し、件数を報告する。
- 出力は同じ入力から常に同じ内容になる。
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import sys
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl

BASE = Path(__file__).resolve().parent.parent
SRC = BASE / "data" / "raw_stations.xlsx"
WATER = BASE / "data" / "raw_water.csv"
TRANSPORT = BASE / "data" / "raw_transport.csv"
SOURCES = BASE / "data" / "sources.json"
CONFIG = BASE / "config.json"
OUT = BASE / "data" / "points.json"
REPORT = BASE / "data" / "build-report.json"

# 東京都本土のおおよその矩形。島しょ部は現行デモの対象外。
TOKYO_BOX = (35.4, 35.95, 138.9, 139.95)

# 東京都の自治体名と住所先頭を最長一致させる。単純な正規表現では
# 「武蔵村山市」を「武蔵村」と誤認するため、自治体名を明示する。
TOKYO_MUNICIPALITIES = tuple(
    sorted(
        """
        千代田区 中央区 港区 新宿区 文京区 台東区 墨田区 江東区 品川区 目黒区
        大田区 世田谷区 渋谷区 中野区 杉並区 豊島区 北区 荒川区 板橋区 練馬区
        足立区 葛飾区 江戸川区 八王子市 立川市 武蔵野市 三鷹市 青梅市 府中市
        昭島市 調布市 町田市 小金井市 小平市 日野市 東村山市 国分寺市 国立市
        福生市 狛江市 東大和市 清瀬市 東久留米市 武蔵村山市 多摩市 稲城市
        羽村市 あきる野市 西東京市 瑞穂町 日の出町 檜原村 奥多摩町 大島町
        利島村 新島村 神津島村 三宅村 御蔵島村 八丈町 青ヶ島村 小笠原村
        """.split(),
        key=len,
        reverse=True,
    )
)

# 既存の短い日本語コードを保ちつつ、実データで衝突した2自治体だけ分ける。
CODE_PREFIX_OVERRIDES = {"武蔵野市": "武野", "武蔵村山市": "武村"}


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def in_tokyo(lat: float, lon: float) -> bool:
    lat_min, lat_max, lon_min, lon_max = TOKYO_BOX
    return lat_min <= lat <= lat_max and lon_min <= lon <= lon_max


def municipality_from(address: str) -> str | None:
    normalized = address.strip()
    if normalized.startswith("東京都"):
        normalized = normalized[3:]
    for municipality in TOKYO_MUNICIPALITIES:
        if normalized.startswith(municipality):
            return municipality
    return None


def code_for(municipality: str, sequence: int) -> str:
    prefix = CODE_PREFIX_OVERRIDES.get(municipality, municipality[:2])
    return f"{prefix}{sequence:04d}"


def coord(value: Any) -> float | None:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def meters(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    earth_radius, radians = 6_371_000, math.pi / 180
    x = (lat2 - lat1) * radians
    y = (lon2 - lon1) * radians * math.cos((lat1 + lat2) / 2 * radians)
    return math.sqrt(x * x + y * y) * earth_radius


def read_cp932_csv(path: Path) -> list[dict[str, str]]:
    body = path.read_bytes().decode("cp932")
    return list(csv.DictReader(io.StringIO(body)))


def build_candidate_points() -> tuple[list[dict[str, Any]], dict[str, int]]:
    workbook = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    expected = ("No", "店舗名", "所在地", "緯度", "経度")
    if len(rows) < 3 or tuple(rows[1][:5]) != expected:
        current = rows[1][:5] if len(rows) > 1 else "missing"
        raise ValueError(f"raw_stations.xlsx の列構成が変わった: {current}")

    sequences: dict[str, int] = {}
    points: list[dict[str, Any]] = []
    skipped = 0
    out_of_box = 0
    unknown_municipality = 0

    for row in rows[2:]:
        if not row or not all(row[index] is not None for index in (1, 2, 3, 4)):
            skipped += 1
            continue
        lat, lon = coord(row[3]), coord(row[4])
        if lat is None or lon is None:
            skipped += 1
            continue
        if not in_tokyo(lat, lon):
            out_of_box += 1
            continue
        address = str(row[2]).strip()
        municipality = municipality_from(address)
        if not municipality:
            unknown_municipality += 1
            continue
        sequences[municipality] = sequences.get(municipality, 0) + 1
        points.append(
            {
                "c": code_for(municipality, sequences[municipality]),
                "n": str(row[1]).strip(),
                "a": address,
                "m": municipality,
                "lat": round(lat, 6),
                "lon": round(lon, 6),
                "status": "candidate",
            }
        )

    duplicate_codes = [code for code, count in Counter(p["c"] for p in points).items() if count > 1]
    if duplicate_codes:
        raise ValueError(f"場所コードが重複した: {duplicate_codes[:10]}")

    return points, {
        "raw_rows_including_two_headers": len(rows),
        "candidate_points": len(points),
        "municipalities": len(sequences),
        "empty_or_invalid_rows": skipped,
        "out_of_tokyo_box_rows": out_of_box,
        "unknown_municipality_rows": unknown_municipality,
    }


def municipality_peers(points: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    peers: dict[str, list[dict[str, Any]]] = {}
    for point in points:
        peers.setdefault(point["m"], []).append(point)
    return peers


def coordinate_matches_municipality(
    item: dict[str, Any], peers: dict[str, list[dict[str, Any]]], maximum_m: float = 3000
) -> bool | None:
    municipality = municipality_from(item.get("a", ""))
    candidates = peers.get(municipality or "")
    if not candidates:
        return None
    nearest = min(
        meters(item["lat"], item["lon"], point["lat"], point["lon"])
        for point in candidates
    )
    return nearest <= maximum_m


def build_destinations(
    points: list[dict[str, Any]], config: dict[str, Any]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    allowed = set(config["data_quality"]["allowed_destination_kinds"])
    max_cluster = int(config["data_quality"]["max_duplicate_destination_coordinate_cluster"])
    peers = municipality_peers(points)
    output: list[dict[str, Any]] = []
    report: dict[str, Any] = {}

    water_rows = read_cp932_csv(WATER)
    water_stats = Counter()
    for row in water_rows:
        water_stats["raw_rows"] += 1
        if (row.get("稼働停止") or "").strip():
            water_stats["inactive"] += 1
            continue
        lat, lon = coord(row.get("緯度")), coord(row.get("経度"))
        if lat is None or lon is None or not in_tokyo(lat, lon):
            water_stats["invalid_coordinate"] += 1
            continue
        item = {
            "k": "water",
            "n": (row.get("施設名") or "").strip(),
            "a": (row.get("所在地") or "").strip(),
            "lat": lat,
            "lon": lon,
            "source_id": "tokyowater-drinking-stations-2026",
        }
        match = coordinate_matches_municipality(item, peers)
        if match is False:
            water_stats["municipality_mismatch"] += 1
            continue
        if match is None:
            water_stats["municipality_unjudged"] += 1
        if "water" in allowed:
            output.append(item)
            water_stats["enabled"] += 1
        else:
            water_stats["quarantined"] += 1
    report["water"] = dict(water_stats)

    transport_rows = read_cp932_csv(TRANSPORT)
    coordinate_counts = Counter(
        ((row.get("緯度") or "").strip(), (row.get("経度") or "").strip())
        for row in transport_rows
    )
    transport_stats = Counter()
    transport_stats["raw_rows"] = len(transport_rows)
    transport_stats["largest_coordinate_cluster"] = max(coordinate_counts.values(), default=0)
    for row in transport_rows:
        lat, lon = coord(row.get("緯度")), coord(row.get("経度"))
        if lat is None or lon is None or not in_tokyo(lat, lon):
            transport_stats["invalid_coordinate"] += 1
            continue
        raw_coordinate = ((row.get("緯度") or "").strip(), (row.get("経度") or "").strip())
        if coordinate_counts[raw_coordinate] > max_cluster:
            transport_stats["duplicate_coordinate_cluster"] += 1
            continue
        item = {
            "k": "station",
            "n": (row.get("施設名") or "").strip(),
            "a": f"{row.get('市区町村名', '')}{row.get('町丁目名', '')}".strip(),
            "ev": (row.get("エレベーターの有無") or "").strip(),
            "lat": lat,
            "lon": lon,
            "source_id": "daredemo-tokyo-transport",
        }
        match = coordinate_matches_municipality(item, peers)
        if match is False:
            transport_stats["municipality_mismatch"] += 1
            continue
        if match is None:
            transport_stats["municipality_unjudged"] += 1
        if "station" in allowed:
            output.append(item)
            transport_stats["enabled"] += 1
        else:
            transport_stats["quarantined"] += 1
    report["transport"] = dict(transport_stats)
    return output, report


def main() -> int:
    required = (SRC, WATER, TRANSPORT, SOURCES, CONFIG)
    missing = [str(path.relative_to(BASE)) for path in required if not path.exists()]
    if missing:
        print(f"✗ 必要ファイルが無い: {', '.join(missing)}", file=sys.stderr)
        return 1

    config = load_json(CONFIG)
    source_manifest = load_json(SOURCES)
    points, point_stats = build_candidate_points()
    destinations, destination_stats = build_destinations(points, config)

    minimum = int(config["data_quality"]["min_candidate_points"])
    maximum = int(config["data_quality"]["max_candidate_points"])
    if not minimum <= len(points) <= maximum:
        raise ValueError(
            f"候補地点数 {len(points)} が許容範囲 {minimum}..{maximum} を外れた。入力更新をレビューすること"
        )

    sources_for_runtime = [
        {
            key: source[key]
            for key in ("id", "title", "provider", "catalog_url", "runtime_status")
        }
        for source in source_manifest["sources"]
    ]
    payload = {
        "schema_version": 1,
        "data_mode": config["data_mode"],
        "note": "候補地を使った実証デモ。実際にステッカーが設置された地点ではない。",
        "capabilities": {
            "waypoint_guidance_public": config["wayfinding"]["enabled_for_public_use"],
            "enabled_destination_kinds": config["data_quality"]["allowed_destination_kinds"],
        },
        "sources": sources_for_runtime,
        "points": points,
        "dest": destinations,
    }
    report = {
        "schema_version": 1,
        "input_sha256": {
            str(path.relative_to(BASE)): sha256(path) for path in (SRC, WATER, TRANSPORT)
        },
        "candidate_points": point_stats,
        "destinations": destination_stats,
        "output": {
            "candidate_points": len(points),
            "destinations": len(destinations),
            "destination_kinds": dict(Counter(item["k"] for item in destinations)),
        },
    }

    OUT.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    REPORT.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print(
        f"✓ 候補地点 {len(points):,} 点 / {point_stats['municipalities']} 自治体 "
        f"（場所コード重複 0）"
    )
    print(
        "   空行・不正値 {empty_or_invalid_rows} ／ 東京都本土の範囲外 {out_of_tokyo_box_rows} "
        "／ 自治体不明 {unknown_municipality_rows}".format(**point_stats)
    )
    print(f"✓ 表示する周辺地点 {len(destinations):,} 点 {report['output']['destination_kinds']}")
    print(
        "⚠ 交通データは隔離: 同一座標クラスター最大 "
        f"{destination_stats['transport'].get('largest_coordinate_cluster', 0)} 件"
    )
    print(f"   → {OUT.relative_to(BASE)} / {REPORT.relative_to(BASE)}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except (KeyError, TypeError, ValueError, json.JSONDecodeError) as error:
        print(f"✗ データ品質ゲート失敗: {error}", file=sys.stderr)
        sys.exit(2)
