#!/usr/bin/env python3
"""4つの公式推計を区市町村名で結合し、画面が読む1本のデータにする。

  python3 scripts/build_dataset.py

出力: data/stress_test.json （画面が読む正本）／data/stress_test.csv （人が見る用）

結合できない自治体は **落とさずに理由つきで excluded に入れる**。
黙って消すと「東京都には53自治体しかない」ように見えてしまう。
"""
from __future__ import annotations

import csv
import io
import json
import sys
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
RAW = BASE / "data" / "raw"
OUT = BASE / "data"

YEARS = [2020, 2025, 2030, 2035, 2040, 2045, 2050]
# 就業者・世帯の推計は 2045 年で切れる。指標はここまでしか出さない（外挿しない）
YEARS_ALL = [2020, 2025, 2030, 2035, 2040, 2045]
TOKYO = "東京都"


def read_csv(name: str, encoding: str) -> list[dict]:
    text = (RAW / name).read_bytes().decode(encoding)
    return list(csv.DictReader(io.StringIO(text)))


def load_population() -> dict[str, dict]:
    """社人研 xlsx。シート1枚＝1自治体。高齢化率と65歳以上は固定行にある。"""
    wb = openpyxl.load_workbook(RAW / "ipss_tokyo_13.xlsx", read_only=True, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(max_row=40, max_col=8, values_only=True))
        rec = {}
        for row in rows:
            label = (str(row[0]) if row[0] is not None else "").strip().replace("　", "")
            values = [row[i + 1] for i in range(7)]
            if label.startswith("年齢別割合（65歳以上"):
                rec["rate"] = values
            elif label == "総数":
                rec["total"] = values
            elif label.startswith("（再掲）65歳以上"):
                rec["elderly"] = values
        if {"rate", "total", "elderly"} <= rec.keys():
            out[sheet_name.split("_", 1)[1]] = rec
    return out


def main() -> int:
    pop = load_population()
    if TOKYO not in pop:
        print("✗ 社人研データに東京都のシートが無い", file=sys.stderr)
        return 1

    emp_rows = read_csv("employment_by_age.csv", "utf-8-sig")
    employment = {
        r["地域名"]: {y: int(r[f"{y}年"]) for y in YEARS_ALL}
        for r in emp_rows
        if r["性別"] == "男女計" and r["年齢階級区分"] == "総数"
    }
    # 65歳以上の昼間就業者＝「老いが価値を生んでいる」側の実数。
    # 凡例は区市町村を階層コード3と書いているが、実データは4（2026-08-23 実測）。
    # ここでは階層コードに頼らず、年齢階級を足し上げる。
    OLD = {"65～69", "70～74", "75～79", "80～84", "85歳以上"}
    elderly_workers: dict[str, dict[int, int]] = {}
    for r in emp_rows:
        if r["性別"] != "男女計" or r["年齢階級区分"] not in OLD:
            continue
        d = elderly_workers.setdefault(r["地域名"], {y: 0 for y in YEARS_ALL})
        for y in YEARS_ALL:
            d[y] += int(r[f"{y}年"])
    general = {r["地域名"]: r for r in read_csv("households_general.csv", "utf-8-sig")}
    single = {
        r["地域名"]: r
        for r in read_csv("households_single.csv", "utf-8-sig")
        if r["世帯主の性別区分"] == "男女計"
    }

    records, excluded = [], []
    for name in [n for n in pop if n != TOKYO]:
        missing = [
            label
            for label, src in (("就業者", employment), ("高齢就業者", elderly_workers),
                               ("一般世帯", general), ("単独世帯", single))
            if name not in src
        ]
        if missing:
            excluded.append({"name": name, "missing": missing})
            continue
        series = {}
        for year in YEARS_ALL:
            i = YEARS.index(year)
            elderly = pop[name]["elderly"][i]
            workers = employment[name][year]
            old_workers = elderly_workers[name][year]
            households = int(general[name][f"{year}年"])
            singles = int(single[name][f"{year}年"])
            series[year] = {
                "aging_rate": round(pop[name]["rate"][i], 1),
                "elderly": elderly,
                "population": pop[name]["total"][i],
                "workers": workers,
                # 支え手比率＝高齢者1人あたりの昼間就業者。3本を結合して初めて出る
                "support_ratio": round(workers / elderly, 2),
                # 高齢者のうち何割が働いているか／就業者のうち何割が高齢者か
                "elderly_workers": old_workers,
                "elderly_working_rate": round(old_workers / elderly * 100, 1),
                "workforce_elderly_share": round(old_workers / workers * 100, 1),
                "single_household_rate": round(singles / households * 100, 1),
            }
        records.append({"name": name, "series": series})

    records.sort(key=lambda r: -r["series"][2045]["elderly_workers"])
    payload = {
        "generated_from": "社人研 地域別将来推計人口(令和5年推計) / 東京都就業者数の予測(令和7年) / 東京都世帯数の予測",
        "years": YEARS_ALL,
        "note": "2020年は実績値、以降は推計。2045年より先へは外挿していない。",
        "municipalities": records,
        "excluded": excluded,
    }
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "stress_test.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")

    with (OUT / "stress_test.csv").open("w", encoding="utf-8-sig", newline="") as f:
        cols = ["自治体"] + [
            f"{k}{y}" for y in (2020, 2030, 2040, 2045)
            for k in ("高齢化率", "高齢者数", "昼間就業者", "65歳以上就業者",
                      "高齢者就業率", "就業者に占める高齢者割合", "支え手比率", "単独世帯率")
        ]
        w = csv.writer(f)
        w.writerow(cols)
        for rec in records:
            row = [rec["name"]]
            for year in (2020, 2030, 2040, 2045):
                s = rec["series"][year]
                row += [s["aging_rate"], s["elderly"], s["workers"], s["elderly_workers"],
                        s["elderly_working_rate"], s["workforce_elderly_share"],
                        s["support_ratio"], s["single_household_rate"]]
            w.writerow(row)

    print(f"✓ {len(records)} 自治体を出力（除外 {len(excluded)} 件）")
    for e in excluded:
        print(f"   除外: {e['name']}（{'・'.join(e['missing'])}が推計に無い）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
