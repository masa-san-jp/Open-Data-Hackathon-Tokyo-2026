#!/usr/bin/env python3
"""data/raw/ の原本を、列名・自治体コード・型を揃えた中間データへ正規化する。

  python3 scripts/normalize_data.py

自治体コードを第一キーにする。2020年は国勢調査の実績値のみを扱う
（2025年以降はIPSSの推計値であり「実測」ではないため、Phase 2 の初版では正規化対象に含めない）。
数値が数値型でない・シートが読めない自治体は、無視せず `excluded` に理由付きで残す。

外部依存: openpyxl（xlsx を読むため。標準ライブラリに xlsx パーサが無いための実務上の依存）。
"""
from __future__ import annotations

import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("normalize_data: openpyxl が無い。`pip install openpyxl` してから実行して", file=sys.stderr)
    sys.exit(3)

BASE = Path(__file__).resolve().parent.parent
RAW_XLSX = BASE / "data" / "raw" / "ipss_tokyo_population.xlsx"
OUT = BASE / "data" / "normalized" / "population.json"

SOURCE_ID = "src-ipss-population-2023"
TARGET_YEAR = 2020  # 国勢調査による実績値の年（推計値ではない）
YEAR_COL = 2  # 2020年の列（B列）

# シート内の固定行（このxlsxの2026-08-23時点の構造。公開元が構造を変えたら止める）
ROW_TOTAL = 5          # 総数
ROW_AGED_65_PLUS = 28  # （再掲）65歳以上
ROW_AGED_SHARE_PCT = 34  # 年齢別割合（65歳以上：％）


def main() -> int:
    if not RAW_XLSX.exists():
        print(f"normalize_data: {RAW_XLSX} が無い。先に fetch_sources.py を実行して", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(RAW_XLSX, data_only=True)

    municipalities = []
    excluded = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        code = ws.cell(row=2, column=1).value
        name = ws.cell(row=2, column=2).value

        if code == 13000:
            continue  # 東京都全体の集計シートは自治体単位の対象外

        total = ws.cell(row=ROW_TOTAL, column=YEAR_COL).value
        aged_n = ws.cell(row=ROW_AGED_65_PLUS, column=YEAR_COL).value
        aged_pct = ws.cell(row=ROW_AGED_SHARE_PCT, column=YEAR_COL).value

        if not isinstance(code, int):
            excluded.append({"sheet": sheet_name, "reason": "自治体コードが数値でない"})
            continue
        if not all(isinstance(v, (int, float)) for v in (total, aged_n, aged_pct)):
            excluded.append({
                "sheet": sheet_name, "code": code, "name": name,
                "reason": "総数・65歳以上人口・高齢化率のいずれかが数値でない（欠測・秘匿の可能性）",
            })
            continue

        municipalities.append({
            "code": code,
            "name": name,
            "year": TARGET_YEAR,
            "population_total": total,
            "population_aged_65_plus": aged_n,
            "aged_share_pct": round(aged_pct, 1),
            "source_id": SOURCE_ID,
        })

    municipalities.sort(key=lambda m: m["code"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": "0.1.0",
        "normalized_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "source_id": SOURCE_ID,
        "target_year": TARGET_YEAR,
        "target_year_basis": "2020年は国勢調査による実績値（IPSS原本シートの表題に明記）",
        "municipality_count": len(municipalities),
        "municipalities": municipalities,
        "excluded": excluded,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"✓ {len(municipalities)} 自治体を正規化、{len(excluded)} 件を除外 -> {OUT}")
    for e in excluded:
        print("   -", e)
    return 0


if __name__ == "__main__":
    sys.exit(main())
