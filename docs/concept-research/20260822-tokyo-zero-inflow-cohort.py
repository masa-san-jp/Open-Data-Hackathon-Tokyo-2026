#!/usr/bin/env python3
"""東京都の国内・国外転入ゼロシナリオを計算するコーホート要因法。

The model is deliberately explicit about the available data.  It models
Japanese and foreign residents separately, and male/female 1-year cohorts.
The Tokyo resident-registry age table is 5-year grouped, so each group is
split uniformly across ages (the 100+ tail uses a small survivorship-shaped
distribution).  Domestic migration is observed by 5-year age, sex and the
two nationality groups.  International migration is observed by sex and
nationality, but not age; its age profile is proxied by the corresponding
domestic-migration profile.

The script downloads the official source files on each run.  It uses only
public URLs and writes a CSV beside the script when invoked with --write-csv.
"""

from __future__ import annotations

import argparse
import csv
import io
import math
import urllib.request
from pathlib import Path
from typing import Iterable

import numpy as np
import openpyxl


BASE_YEAR = 2026
LAST_YEAR = 2100
MAX_AGE = 110
NAT_NAMES = ("日本人", "外国人")
SEX_NAMES = ("男", "女")
TOKYO_CODE = "13000"
TOKYO_POPULATION_ESTIMATE_2026 = 14_270_748

REGISTRY_URL = "https://www.toukei.metro.tokyo.lg.jp/juukiy/2026/jy26qv0700.csv"
TOKYO_ANNUAL_URL = "https://www.toukei.metro.tokyo.lg.jp/jugoki/2025/ju25qv0100.csv"
TOKYO_INFLOW_URL = "https://www.toukei.metro.tokyo.lg.jp/jugoki/2025/ju25qv0900.csv"
TOKYO_OUTFLOW_URL = "https://www.toukei.metro.tokyo.lg.jp/jugoki/2025/ju25qv1000.csv"
TOKYO_BIRTH_URL = "https://www.toukei.metro.tokyo.lg.jp/jugoki/2025/ju25qv1500.csv"

DOMESTIC_INFLOW_URL = (
    "https://www.e-stat.go.jp/stat-search/file-download?"
    "statInfId=000040407040&fileKind=0"
)
DOMESTIC_OUTFLOW_URL = (
    "https://www.e-stat.go.jp/stat-search/file-download?"
    "statInfId=000040407041&fileKind=0"
)
INTERNATIONAL_URL = (
    "https://www.e-stat.go.jp/stat-search/file-download?"
    "statInfId=000040407055&fileKind=0"
)
IPSS_LIFE_SCHEDULE_URL = (
    "https://www.ipss.go.jp/pp-zenkoku/j/zenkoku2023/db_zenkoku2023/"
    "s_tables/ATable7.xlsx"
)


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "tokyo-cohort-model/1.0"})
    with urllib.request.urlopen(request, timeout=90) as response:
        return response.read()


def read_csv_url(url: str) -> list[dict[str, str]]:
    text = fetch(url).decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def num(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").strip()
        if value in ("", "-", "…", "－"):
            return 0.0
    return float(value)


def new_state() -> np.ndarray:
    return np.zeros((2, 2, MAX_AGE + 1), dtype=float)


def age_group_ranges() -> list[range]:
    return [range(start, start + 5) for start in range(0, 90, 5)] + [range(90, MAX_AGE + 1)]


def tail_weights() -> np.ndarray:
    # A small approximation for the 100+ open-ended registry category.
    weights = np.array(
        [0.51, 0.25, 0.12, 0.06, 0.03, 0.015, 0.007, 0.004, 0.002, 0.001, 0.0],
        dtype=float,
    )
    return weights / weights.sum()


def load_registry_population() -> np.ndarray:
    rows = read_csv_url(REGISTRY_URL)
    state = new_state()
    ranges = age_group_ranges()
    for row in rows:
        if row.get("地域コード") != TOKYO_CODE:
            continue
        age_label = row.get("年齢", "")
        if age_label == "総数" or age_label == "不詳者":
            continue
        if "以上" in age_label:
            age_range = range(100, MAX_AGE + 1)
        else:
            try:
                start = int(age_label.split("～")[0])
            except (ValueError, IndexError):
                continue
            age_range = range(start, min(start + 5, MAX_AGE + 1))
        for nat_index, nat_prefix in enumerate(("日本人", "外国人")):
            for sex_index, sex_prefix in enumerate(("男", "女")):
                column = f"{nat_prefix}／{sex_prefix}(人)"
                value = num(row.get(column))
                if age_label == "100以上":
                    for age, weight in zip(age_range, tail_weights()):
                        state[nat_index, sex_index, age] += value * weight
                else:
                    for age in age_range:
                        state[nat_index, sex_index, age] += value / len(age_range)

    # The registry age table and Tokyo's estimated population have different
    # bases.  Preserve the detailed composition, but calibrate the sum to the
    # official 1 Jan 2026 Tokyo estimate.
    total = state.sum()
    if total <= 0:
        raise RuntimeError("Tokyo registry population was not loaded")
    state *= TOKYO_POPULATION_ESTIMATE_2026 / total
    return state


def tokyo_national_totals(url: str, value_suffix: str) -> dict[str, float]:
    rows = read_csv_url(url)
    output: dict[str, float] = {}
    for row in rows:
        if row.get("地域コード") != TOKYO_CODE:
            continue
        category = row.get("国籍区分", "総数")
        output[category] = num(row.get(value_suffix))
    return output


def tokyo_2025_components() -> dict[str, float]:
    rows = read_csv_url(TOKYO_ANNUAL_URL)
    row = next(r for r in rows if r.get("地域コード") == TOKYO_CODE)
    return {
        "population_2026": num(row["令和8年1月1日現在人口"]),
        "population_change_2025": num(row["令和7年中の動き／人口増減"]),
        "births": num(row["令和7年中の動き／自然増減／出生数"]),
        "deaths": num(row["令和7年中の動き／自然増減／死亡数"]),
        "other_pref_inflow": num(row["令和7年中の動き／社会増減（他県との移動増減）／転入数"]),
        "other_pref_outflow": num(row["令和7年中の動き／社会増減（他県との移動増減）／転出数"]),
    }


def split_flow_to_ages(values: Iterable[float], base: np.ndarray, nat: int, sex: int) -> np.ndarray:
    result = np.zeros(MAX_AGE + 1, dtype=float)
    for value, age_range in zip(values, age_group_ranges()):
        value = float(value)
        if age_range.start < 90:
            for age in age_range:
                result[age] += value / len(age_range)
        else:
            weights = base[nat, sex, 90:] / base[nat, sex, 90:].sum()
            result[90:] += value * weights
    return result


def load_domestic_flow(url: str, base: np.ndarray, target_totals: dict[str, float]) -> np.ndarray:
    workbook = openpyxl.load_workbook(io.BytesIO(fetch(url)), data_only=True, read_only=True)
    worksheet = workbook.active
    output = new_state()
    nat_map = {"日本人移動者": 0, "外国人移動者": 1}
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 67 or row[6] != "東京都":
            continue
        nat_name = row[2]
        if nat_name not in nat_map:
            continue
        nat = nat_map[nat_name]
        total_values = [num(v) for v in row[8:27]]
        male_values = [num(v) for v in row[28:47]]
        female_values = [num(v) for v in row[48:67]]
        output[nat, 0] = split_flow_to_ages(male_values, base, nat, 0)
        output[nat, 1] = split_flow_to_ages(female_values, base, nat, 1)
        # The published total is kept for a validation check; the sex rows are
        # the actual age/sex input.
        if not math.isclose(sum(total_values), num(row[7]), rel_tol=0, abs_tol=2):
            raise RuntimeError(f"Domestic flow total mismatch for {nat_name}")

    # Align the age/sex table to Tokyo's local annual total.  The difference is
    # small and comes from the statistical table's annual aggregation rules.
    for nat, nat_name in enumerate(NAT_NAMES):
        actual = output[nat].sum()
        target = target_totals.get(nat_name, actual)
        if actual > 0:
            output[nat] *= target / actual
    return output


def load_international_flows(base: np.ndarray, domestic_in: np.ndarray, domestic_out: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    workbook = openpyxl.load_workbook(io.BytesIO(fetch(INTERNATIONAL_URL)), data_only=True, read_only=True)
    worksheet = workbook.active
    inflow = new_state()
    outflow = new_state()
    for row in worksheet.iter_rows(values_only=True):
        if len(row) < 12 or row[5] != "東京都":
            continue
        nat_name = row[1]
        if nat_name not in NAT_NAMES:
            continue
        nat = NAT_NAMES.index(nat_name)
        sex_counts_in = [num(row[7]), num(row[8])]
        sex_counts_out = [num(row[10]), num(row[11])]
        for sex in range(2):
            in_profile = domestic_in[nat, sex]
            out_profile = domestic_out[nat, sex]
            if in_profile.sum() == 0:
                in_profile = base[nat, sex]
            if out_profile.sum() == 0:
                out_profile = base[nat, sex]
            inflow[nat, sex] = sex_counts_in[sex] * in_profile / in_profile.sum()
            outflow[nat, sex] = sex_counts_out[sex] * out_profile / out_profile.sum()
    if inflow.sum() <= 0 or outflow.sum() <= 0:
        raise RuntimeError("International movement rows for Tokyo were not loaded")
    return inflow, outflow


def load_mortality() -> np.ndarray:
    workbook = openpyxl.load_workbook(
        io.BytesIO(fetch(IPSS_LIFE_SCHEDULE_URL)), data_only=True, read_only=True
    )
    # ATable7 contains the central mortality assumption for 2020--2070,
    # one male and one female sheet per year.  After 2070, the 2070 schedule
    # is held constant for this long-range Tokyo stress test.
    schedule = np.ones((51, 2, MAX_AGE + 1), dtype=float)
    for year_index in range(51):
        for sex, sex_letter in enumerate(("M", "F")):
            worksheet = workbook[f"AT7({year_index + 1}){sex_letter}"]
            for row in worksheet.iter_rows(values_only=True):
                for age_column, qx_column in ((0, 1), (6, 7)):
                    if len(row) <= qx_column:
                        continue
                    age = row[age_column]
                    rate = row[qx_column]
                    if isinstance(age, (int, float)) and isinstance(rate, (int, float)):
                        age = int(age)
                        if 0 <= age <= MAX_AGE:
                            schedule[year_index, sex, age] = min(max(float(rate), 0.0), 1.0)
    schedule[:, :, 106:] = 1.0
    return schedule


def fertility_rates(base: np.ndarray, births_by_nat: dict[str, float]) -> np.ndarray:
    # Tokyo-like age schedule.  It is scaled independently for Japanese and
    # foreign mothers so that 2025 births match the Tokyo official total.
    schedule = np.zeros(MAX_AGE + 1, dtype=float)
    five_year_rates = {
        15: 0.001,
        20: 0.012,
        25: 0.045,
        30: 0.065,
        35: 0.050,
        40: 0.018,
        45: 0.003,
    }
    for start, rate in five_year_rates.items():
        schedule[start : start + 5] = rate
    output = np.zeros((2, MAX_AGE + 1), dtype=float)
    for nat, nat_name in enumerate(NAT_NAMES):
        unscaled = float(np.dot(base[nat, 1], schedule))
        target = births_by_nat.get(nat_name, 0.0)
        if unscaled <= 0:
            raise RuntimeError(f"No women of reproductive age for {nat_name}")
        output[nat] = schedule * target / unscaled
    return output


def mortality_for_year(qx: np.ndarray, year: int, nat: int, sex: int) -> np.ndarray:
    del nat  # Tokyo does not publish a stable foreign-specific schedule.
    schedule_index = min(max(year - 2020, 0), 50)
    return qx[schedule_index, sex].copy()


def make_inputs() -> dict[str, object]:
    base = load_registry_population()
    inflow_totals = tokyo_national_totals(TOKYO_INFLOW_URL, "令和7年中の転入者数")
    outflow_totals = tokyo_national_totals(TOKYO_OUTFLOW_URL, "令和7年中の転出者数")
    births = tokyo_national_totals(TOKYO_BIRTH_URL, "令和7年中の出生数")
    domestic_in = load_domestic_flow(DOMESTIC_INFLOW_URL, base, inflow_totals)
    domestic_out = load_domestic_flow(DOMESTIC_OUTFLOW_URL, base, outflow_totals)
    international_in, international_out = load_international_flows(base, domestic_in, domestic_out)
    return {
        "base": base,
        "domestic_in": domestic_in,
        "domestic_out": domestic_out,
        "international_in": international_in,
        "international_out": international_out,
        "mortality": load_mortality(),
        "fertility": fertility_rates(base, births),
        "components": tokyo_2025_components(),
        "births_by_nat": births,
        "inflow_totals": inflow_totals,
        "outflow_totals": outflow_totals,
    }


def summary(year: int, state: np.ndarray) -> dict[str, float]:
    total = float(state.sum())
    japanese = float(state[0].sum())
    foreign = float(state[1].sum())
    under15 = float(state[:, :, :15].sum())
    working = float(state[:, :, 15:65].sum())
    elderly = float(state[:, :, 65:].sum())
    return {
        "year": year,
        "total": total,
        "japanese": japanese,
        "foreign": foreign,
        "male": float(state[:, 0].sum()),
        "female": float(state[:, 1].sum()),
        "under15": under15,
        "working_age_15_64": working,
        "age65plus": elderly,
        "foreign_share": foreign / total if total else 0.0,
        "age65plus_share": elderly / total if total else 0.0,
    }


def project(inputs: dict[str, object], domestic_inflow: bool, international_inflow: bool) -> list[dict[str, float]]:
    state = np.asarray(inputs["base"], dtype=float).copy()
    domestic_in = np.asarray(inputs["domestic_in"], dtype=float)
    domestic_out = np.asarray(inputs["domestic_out"], dtype=float)
    international_in = np.asarray(inputs["international_in"], dtype=float)
    international_out = np.asarray(inputs["international_out"], dtype=float)
    qx = np.asarray(inputs["mortality"], dtype=float)
    fertility = np.asarray(inputs["fertility"], dtype=float)

    # Outflow is expressed as a 2025 age/sex/nationality rate, so it declines
    # with the cohort base.  Gross inflow, when enabled for a comparison run,
    # is held at the observed 2025 count rather than being multiplied by the
    # shrinking/expanding Tokyo stock.
    base = np.asarray(inputs["base"], dtype=float)
    dom_out_rate = np.divide(domestic_out, base, out=np.zeros_like(base), where=base > 0)
    int_out_rate = np.divide(international_out, base, out=np.zeros_like(base), where=base > 0)

    records = [summary(BASE_YEAR, state)]
    for year in range(BASE_YEAR, LAST_YEAR):
        remaining = np.zeros_like(state)
        for nat in range(2):
            # The two sex schedules are applied separately.
            for sex in range(2):
                mortality = mortality_for_year(qx, year, nat, sex)
                survivors = state[nat, sex] * (1.0 - mortality)
                survivors *= np.maximum(0.0, 1.0 - dom_out_rate[nat, sex])
                survivors *= np.maximum(0.0, 1.0 - int_out_rate[nat, sex])
                remaining[nat, sex, 1:] = survivors[:-1]

        births = np.zeros(2, dtype=float)
        for nat in range(2):
            births[nat] = float(np.dot(state[nat, 1], fertility[nat]))
            # The mother nationality is used as the cohort nationality.
            male_share = 0.512
            remaining[nat, 0, 0] += births[nat] * male_share
            remaining[nat, 1, 0] += births[nat] * (1.0 - male_share)

        if domestic_inflow:
            remaining += domestic_in
        if international_inflow:
            remaining += international_in

        state = remaining
        records.append(summary(year + 1, state))
    return records


def format_million(value: float) -> str:
    return f"{value / 1_000_000:.3f}"


def print_results(inputs: dict[str, object], results: dict[str, list[dict[str, float]]]) -> None:
    years = [2026, 2030, 2050, 2075, 2100]
    print("scenario,year,total_million,japanese_million,foreign_million,age65plus_share")
    for scenario, rows in results.items():
        by_year = {int(row["year"]): row for row in rows}
        for year in years:
            row = by_year[year]
            print(
                f"{scenario},{year},{format_million(row['total'])},"
                f"{format_million(row['japanese'])},{format_million(row['foreign'])},"
                f"{row['age65plus_share']:.3f}"
            )
    components = inputs["components"]
    print("\n2025 Tokyo observed components")
    for key, value in components.items():
        print(f"{key}: {value:,.0f}")


def write_csv(path: Path, results: dict[str, list[dict[str, float]]]) -> None:
    rows = [dict({"scenario": scenario}, **row) for scenario, records in results.items() for row in records]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write-csv", type=Path, help="write all annual scenario rows to this CSV")
    args = parser.parse_args()

    inputs = make_inputs()
    scenarios = {
        "both_inflows_zero": (False, False),
        "domestic_inflow_zero": (False, True),
        "international_inflow_zero": (True, False),
        "observed_inflows_reference": (True, True),
    }
    results = {
        name: project(inputs, domestic_inflow=flags[0], international_inflow=flags[1])
        for name, flags in scenarios.items()
    }
    print_results(inputs, results)
    if args.write_csv:
        write_csv(args.write_csv, results)


if __name__ == "__main__":
    main()
